"""
excel_reader.py

負責：
- 使用 openpyxl 讀取 Excel 檔案
- 將第一列視為表頭
- 將每一列資料轉為 dict
- 自動保留 row_number，方便錯誤追蹤與輸出命名
- 忽略完全空白的資料列

業務語意：
- source = 對應 Excel 欄位名稱
- output_name = 最終輸出檔名
"""

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


def read_excel_rows(excel_path):
    """
    讀取 Excel 檔案並將每一列資料轉換為 dict。

    流程：
    1. 使用 openpyxl 載入 workbook
    2. 第一列作為 header
    3. 每列資料轉為 dict
    4. 自動加入 row_number 方便錯誤追蹤
    5. 忽略完全空白的列

    參數：
    excel_path (Path | str)
        Excel 檔案路徑

    回傳：
    list[dict]
        每一列資料所組成的清單
    """
    excel_path = Path(excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"找不到 Excel 檔案：{excel_path}")

    try:
        workbook = load_workbook(excel_path, data_only=True)
    except InvalidFileException as exc:
        raise ValueError(
            f"Excel 檔案格式不合法，請確認是可讀取的 .xlsx 檔案：{excel_path}"
        ) from exc

    try:
        worksheet = workbook.active

        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = rows[0]
        if not headers or all(header is None for header in headers):
            raise ValueError("Excel 的第一列必須是表頭。")

        normalized_headers = []
        for column_index, header in enumerate(headers, start=1):
            if header is None or str(header).strip() == "":
                raise ValueError(
                    f"第 {column_index} 欄的表頭為空白，所有表頭都必須有名稱。"
                )
            # 將表頭統一轉為去除前後空白的字串，避免查欄位時受到格式差異影響。
            normalized_headers.append(str(header).strip())

        data_rows = []
        for excel_row_number, values in enumerate(rows[1:], start=2):
            # 忽略完全沒有有效內容的列，避免產出空資料造成後續錯誤。
            if values is None or all(value is None or str(value).strip() == "" for value in values):
                continue

            # 使用表頭與列值組成 dict，讓後續程式可以用欄位名稱取值。
            row_data = dict(zip(normalized_headers, values))
            # 保留 Excel 的實際列號，方便輸出檔名與錯誤訊息定位。
            row_data["row_number"] = excel_row_number
            data_rows.append(row_data)

        return data_rows
    finally:
        workbook.close()
